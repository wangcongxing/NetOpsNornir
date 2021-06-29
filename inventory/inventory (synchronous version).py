from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side
from netmiko import ConnectHandler
from pprint import pprint

connection_fail = []

wb = Workbook()
ws = wb.active
ws.title = 'Inventory'
ws['A1'] = 'Hostname'
ws['B1'] = 'IP Address'
ws['C1'] = 'Serial Number'
ws['D1'] = 'Uptime'
ws['E1'] = 'Model'
ws['F1'] = 'OS Version'
yellowFill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
ws['A1'].fill=yellowFill
ws['B1'].fill=yellowFill
ws['C1'].fill=yellowFill
ws['D1'].fill=yellowFill
ws['E1'].fill=yellowFill
ws['F1'].fill=yellowFill

with open('ip_list.txt') as f:
    number_of_sw = len(f.readlines()) + 2
    f.seek(0)
    for ips, row in zip(f.readlines(), range(2, number_of_sw)):
        try:
            ip = ips.strip()
            connection_info = {
                    'device_type': 'cisco_ios',
                    'ip': ip,
                    'username': 'parry',
                    'password': 'afmo9se8e!'
            }
            with ConnectHandler(**connection_info) as conn:
                print (f'已经成功登陆交换机{ip}')
                hostname = conn.find_prompt().replace('#','')
                output = conn.send_command('show version', use_textfsm=True)
                sn = output[0]['serial'][0]
                uptime = output[0]['uptime']
                model = output[0]['hardware'][0]
                os_version = output[0]['version']
                ws.cell(row=row, column=1, value=hostname)
                ws.cell(row=row, column=2, value=ip)
                ws.cell(row=row, column=3, value=sn)
                ws.cell(row=row, column=4, value=uptime)
                ws.cell(row=row, column=5, value=model)
                ws.cell(row=row, column=6, value=os_version)
        except Exception as e:
            print (e)
            connection_fail.append(ip)

dims = {}
for row in ws.rows:
    for cell in row:
        cell.border=thin_border
        if cell.value:
            dims[cell.column_letter] = max((dims.get(cell.column, 0), len(str(cell.value))))

for col, value in dims.items():
    ws.column_dimensions[col].width = value + 5

wb.save('inventory.xlsx')

print ('SSH connection to below switches failed: ')
for i in connection_fail:
    print (i)